from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from docx import Document
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

from .validation import MIME_DOCX, MIME_DWG, MIME_PDF, MIME_XLSX, FileValidationError


@dataclass(frozen=True)
class ExtractedUnit:
    ordinal: int
    locator_type: str
    content: str
    content_hash: str
    page_number: int | None = None
    paragraph_number: int | None = None
    sheet_name: str | None = None
    cell_range: str | None = None


@dataclass(frozen=True)
class ExtractionResult:
    units: list[ExtractedUnit]
    metrics: dict[str, Any]
    preview_strategy: str


def extract_document(path: Path, mime_type: str, *, max_units: int = 20_000) -> ExtractionResult:
    if mime_type == MIME_PDF:
        return _extract_pdf(path, max_units)
    if mime_type == MIME_DOCX:
        return _extract_docx(path, max_units)
    if mime_type == MIME_XLSX:
        return _extract_xlsx(path, max_units)
    if mime_type == MIME_DWG:
        return ExtractionResult([], {"unit_count": 0, "mode": "validation_only"}, "cad_adapter_required")
    raise FileValidationError("UNSUPPORTED_EXTRACTION", "No extraction adapter is available for this file type.")


def _extract_pdf(path: Path, max_units: int) -> ExtractionResult:
    reader = PdfReader(path)
    if reader.is_encrypted:
        raise FileValidationError("ENCRYPTED_PDF", "Encrypted PDF files cannot be processed.")
    if len(reader.pages) > max_units:
        raise FileValidationError("PAGE_LIMIT", "PDF exceeds the processing page limit.")
    units = [_unit(index, "page", page.extract_text() or "", page_number=index + 1)
             for index, page in enumerate(reader.pages)]
    return ExtractionResult(units, {"page_count": len(units), "unit_count": len(units)}, "original_pdf")


def _extract_docx(path: Path, max_units: int) -> ExtractionResult:
    document = Document(str(path))
    units: list[ExtractedUnit] = []
    for paragraph_number, paragraph in enumerate(document.paragraphs, start=1):
        content = paragraph.text.strip()
        if not content:
            continue
        if len(units) >= max_units:
            raise FileValidationError("UNIT_LIMIT", "DOCX exceeds the extraction unit limit.")
        units.append(_unit(len(units), "paragraph", content, paragraph_number=paragraph_number))
    return ExtractionResult(units, {"paragraph_count": len(units), "unit_count": len(units)}, "rendered_pdf_pending")


def _extract_xlsx(path: Path, max_units: int) -> ExtractionResult:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    units: list[ExtractedUnit] = []
    visited_cells = 0
    try:
        for sheet in workbook.worksheets:
            for row_number, row in enumerate(sheet.iter_rows(), start=1):
                values = [(index, cell.value) for index, cell in enumerate(row, start=1) if cell.value is not None]
                visited_cells += len(row)
                if visited_cells > 1_000_000:
                    raise FileValidationError("CELL_LIMIT", "XLSX exceeds the processing cell limit.")
                if not values:
                    continue
                if len(units) >= max_units:
                    raise FileValidationError("UNIT_LIMIT", "XLSX exceeds the extraction unit limit.")
                start_col, end_col = values[0][0], values[-1][0]
                cell_range = f"{get_column_letter(start_col)}{row_number}:{get_column_letter(end_col)}{row_number}"
                content = " | ".join(f"{get_column_letter(column)}{row_number}: {value}" for column, value in values)
                units.append(_unit(len(units), "sheet_range", content, sheet_name=sheet.title, cell_range=cell_range))
    finally:
        workbook.close()
    return ExtractionResult(units, {"sheet_count": len(workbook.sheetnames), "unit_count": len(units)}, "rendered_pdf_pending")


def _unit(ordinal: int, locator_type: str, content: str, **locator: Any) -> ExtractedUnit:
    normalised = content.replace("\x00", "").strip()
    return ExtractedUnit(ordinal=ordinal, locator_type=locator_type, content=normalised,
                         content_hash=hashlib.sha256(normalised.encode("utf-8")).hexdigest(), **locator)
