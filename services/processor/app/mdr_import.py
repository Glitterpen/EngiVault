from __future__ import annotations

import io
import re
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import PurePath

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

MAX_IMPORT_ROWS = 500
MAX_ARCHIVE_BYTES = 50 * 1024 * 1024
MAX_ARCHIVE_ENTRIES = 2_000

HEADER_ALIASES = {
    "document number": "document_number",
    "document no": "document_number",
    "doc number": "document_number",
    "doc no": "document_number",
    "title": "title",
    "document title": "title",
    "discipline": "discipline",
    "document type": "document_type",
    "doc type": "document_type",
    "type": "document_type",
    "planned submission date": "planned_submission_date",
    "agreed submission date": "planned_submission_date",
    "first submission date": "planned_submission_date",
    "planned final date": "planned_final_date",
    "final date": "planned_final_date",
    "required issue status": "required_issue_status",
    "required final issue status": "required_issue_status",
    "responsible party": "responsible_party",
    "responsible engineer": "responsible_party",
    "originator": "responsible_party",
    "progress weight": "progress_weight",
    "weight": "progress_weight",
    "area": "area",
    "system": "system",
    "work package": "work_package",
    "workpackage": "work_package",
}
REQUIRED_HEADERS = {
    "document_number",
    "title",
    "discipline",
    "document_type",
    "planned_submission_date",
}


class MdrImportError(ValueError):
    pass


def parse_mdr_workbook(content: bytes, filename: str) -> dict[str, object]:
    if PurePath(filename).suffix.lower() != ".xlsx":
        raise MdrImportError("Upload an Excel .xlsx workbook.")
    _validate_archive(content)
    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as error:
        raise MdrImportError("The Excel workbook could not be opened.") from error

    try:
        worksheet = _select_worksheet(workbook)
        header_row, columns = _find_header(worksheet)
        rows = _read_rows(worksheet, header_row, columns)
        if not rows:
            raise MdrImportError("The workbook contains headings but no document rows.")
        return {
            "sheet_name": worksheet.title,
            "header_row": header_row,
            "rows": rows,
            "row_count": len(rows),
        }
    finally:
        workbook.close()


def _validate_archive(content: bytes) -> None:
    if len(content) < 4 or not content.startswith(b"PK"):
        raise MdrImportError("The uploaded file is not a valid .xlsx workbook.")
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ARCHIVE_ENTRIES:
                raise MdrImportError("The workbook contains too many internal files.")
            if sum(entry.file_size for entry in entries) > MAX_ARCHIVE_BYTES:
                raise MdrImportError("The expanded workbook is too large to import safely.")
            if "[Content_Types].xml" not in archive.namelist():
                raise MdrImportError("The uploaded file is not a valid .xlsx workbook.")
    except zipfile.BadZipFile as error:
        raise MdrImportError("The uploaded file is not a valid .xlsx workbook.") from error


def _select_worksheet(workbook):
    for worksheet in workbook.worksheets:
        if worksheet.title.strip().lower() == "mdr import":
            return worksheet
    return workbook.worksheets[0]


def _find_header(worksheet) -> tuple[int, dict[int, str]]:
    maximum_header_row = min(20, worksheet.max_row) if worksheet.max_row else 20
    for row in worksheet.iter_rows(min_row=1, max_row=maximum_header_row):
        columns: dict[int, str] = {}
        for cell in row:
            if cell.data_type == "f":
                continue
            canonical = HEADER_ALIASES.get(_normalise_header(cell.value))
            if canonical:
                if canonical in columns.values():
                    raise MdrImportError(f"The heading '{cell.value}' appears more than once.")
                columns[cell.column] = canonical
        if REQUIRED_HEADERS.issubset(columns.values()):
            return row[0].row, columns
    required = "Document Number, Title, Discipline, Document Type and Planned Submission Date"
    raise MdrImportError(f"Required headings were not found. Include: {required}.")


def _read_rows(worksheet, header_row: int, columns: dict[int, str]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for cells in worksheet.iter_rows(min_row=header_row + 1, max_col=max(columns)):
        if all(_empty(cells[column - 1].value) for column in columns):
            continue
        if len(rows) >= MAX_IMPORT_ROWS:
            raise MdrImportError(f"A maximum of {MAX_IMPORT_ROWS} document rows can be imported at once.")
        row: dict[str, object] = {"row_number": cells[0].row, "errors": []}
        for column, canonical in columns.items():
            cell = cells[column - 1]
            if cell.data_type == "f":
                row[canonical] = None
                row["errors"].append(f"{_label(canonical)} cannot contain a formula.")
                continue
            row[canonical] = _convert_value(canonical, cell)
        _validate_row(row)
        rows.append(row)
    return rows


def _convert_value(canonical: str, cell: Cell) -> object:
    value = cell.value
    if _empty(value):
        return None
    if canonical in {"planned_submission_date", "planned_final_date"}:
        return _date_value(value)
    if canonical == "progress_weight":
        try:
            return float(Decimal(str(value).strip()))
        except (InvalidOperation, ValueError):
            return "INVALID_WEIGHT"
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value)).strip()


def _validate_row(row: dict[str, object]) -> None:
    errors = row["errors"]
    for required in REQUIRED_HEADERS:
        if _empty(row.get(required)):
            errors.append(f"{_label(required)} is required.")
    if row.get("planned_submission_date") == "INVALID_DATE":
        errors.append("Planned Submission Date must be an Excel date or YYYY-MM-DD.")
        row["planned_submission_date"] = None
    if row.get("planned_final_date") == "INVALID_DATE":
        errors.append("Planned Final Date must be an Excel date or YYYY-MM-DD.")
        row["planned_final_date"] = None
    weight = row.get("progress_weight")
    if weight == "INVALID_WEIGHT":
        errors.append("Progress Weight must be numeric.")
        row["progress_weight"] = None
        weight = None
    if weight is not None and (not isinstance(weight, (int, float)) or weight <= 0 or weight > 1000):
        errors.append("Progress Weight must be greater than 0 and no more than 1000.")
        row["progress_weight"] = None
    submission = row.get("planned_submission_date")
    final = row.get("planned_final_date")
    if isinstance(submission, str) and isinstance(final, str) and final < submission:
        errors.append("Planned Final Date cannot be before Planned Submission Date.")


def _date_value(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(raw, pattern).date().isoformat()  # noqa: DTZ007
        except ValueError:
            continue
    return "INVALID_DATE"


def _normalise_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _label(canonical: str) -> str:
    return canonical.replace("_", " ").title()


def _empty(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())
